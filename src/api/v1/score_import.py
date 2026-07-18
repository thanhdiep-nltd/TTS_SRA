from io import BytesIO
from urllib.parse import quote
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from src.api.deps import get_current_user, get_db
from src.models import enums, tables
from src.schemas.score import ScoreImportConfirmRequest
from src.services import rbac
from src.services.scoring import SCORE_COLUMNS, column_label

router = APIRouter(prefix="/scores", tags=["Score Import"])


@router.get("/import/template")
def download_template(
    class_id: UUID = Query(...),
    subject_id: UUID = Query(...),
    semester_id: UUID = Query(...),
    db: Session = Depends(get_db),
    current_user: tables.User = Depends(get_current_user),
):
    """Sinh file mẫu Excel (.xlsx) chứa sẵn thông tin học sinh và các cột điểm để nhập."""
    # Kiểm tra quyền ghi điểm của user
    if not rbac.can_write_score(db, current_user, subject_id, class_id):
        raise HTTPException(
            status_code=403, detail="Bạn không có quyền nhập hoặc chỉnh sửa điểm số cho lớp học và môn học này."
        )

    # Lấy thông tin lớp, môn học, học kỳ
    cls = db.get(tables.Class, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail="Lớp học không tồn tại")
    subj = db.get(tables.Subject, subject_id)
    if not subj:
        raise HTTPException(status_code=404, detail="Môn học không tồn tại")
    sem = db.get(tables.Semester, semester_id)
    if not sem:
        raise HTTPException(status_code=404, detail="Học kỳ không tồn tại")

    # Lấy danh sách học sinh đang ghi danh trong lớp
    student_stmt = (
        select(tables.Student)
        .join(tables.Enrollment)
        .where(
            tables.Enrollment.class_id == class_id,
            tables.Enrollment.is_active.is_(True),
            tables.Student.is_active.is_(True),
        )
        .order_by(tables.Student.full_name)
    )
    students = db.execute(student_stmt).scalars().all()

    # Lấy các điểm số hiện tại đã có trong hệ thống
    score_stmt = select(tables.Score).where(
        tables.Score.class_id == class_id,
        tables.Score.subject_id == subject_id,
        tables.Score.semester_id == semester_id,
    )
    existing_scores = {}
    for s in db.execute(score_stmt).scalars().all():
        existing_scores[(s.student_id, s.score_category, s.column_index)] = float(s.value)

    # Khởi tạo Workbook openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng nhập điểm"

    # Định nghĩa định dạng (style) cho file mẫu
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="0D4D8B", end_color="0D4D8B", fill_type="solid")
    fill_hidden = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_hidden = Font(name="Arial", size=9, color="94A3B8")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Tiêu đề cột
    headers = ["ID Học sinh (Ẩn)", "Mã học sinh", "Họ và tên"]
    for cat, idx in SCORE_COLUMNS:
        headers.append(column_label(cat, idx))

    ws.append(headers)

    # Định dạng dòng tiêu đề (Header Row)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        if col_idx == 1:
            # Đổi màu riêng cho cột ID Học sinh để nhận biết là cột kỹ thuật ẩn
            cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        else:
            cell.fill = fill_header

    # Ghi dữ liệu từng dòng học sinh
    for r_idx, student in enumerate(students, 2):
        ws.cell(row=r_idx, column=1, value=str(student.id))
        ws.cell(row=r_idx, column=2, value=student.student_code)
        ws.cell(row=r_idx, column=3, value=student.full_name)

        # Định dạng cột ID (kỹ thuật ẩn)
        c1 = ws.cell(row=r_idx, column=1)
        c1.font = font_hidden
        c1.fill = fill_hidden
        c1.alignment = align_center
        c1.border = thin_border

        # Định dạng cột Mã HS
        c2 = ws.cell(row=r_idx, column=2)
        c2.alignment = align_center
        c2.border = thin_border

        # Định dạng cột Tên
        c3 = ws.cell(row=r_idx, column=3)
        c3.alignment = align_left
        c3.border = thin_border

        # Ghi các điểm số hiện có vào các cột điểm tương ứng
        for c_offset, (cat, idx) in enumerate(SCORE_COLUMNS, 4):
            val = existing_scores.get((student.id, cat, idx), "")
            cell = ws.cell(row=r_idx, column=c_offset, value=val)
            cell.alignment = align_center
            cell.border = thin_border

    # Tự động điều chỉnh độ rộng cột vừa vặn nội dung
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Ẩn cột A (ID Học sinh) để tránh giáo viên tá hỏa khi nhìn thấy chuỗi UUID
    ws.column_dimensions["A"].hidden = True

    # Lưu tệp tin vào bộ nhớ đệm BytesIO
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"Mau_nhap_diem_Lop_{cls.name}_Mon_{subj.name}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.post("/import/preview")
async def preview_import(
    class_id: UUID = Form(...),
    subject_id: UUID = Form(...),
    semester_id: UUID = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: tables.User = Depends(get_current_user),
):
    """Đọc file Excel tải lên, kiểm tra tính hợp lệ dữ liệu và trả về kết quả xem trước kèm cảnh báo lỗi."""
    if not rbac.can_write_score(db, current_user, subject_id, class_id):
        raise HTTPException(
            status_code=403, detail="Bạn không có quyền nhập hoặc chỉnh sửa điểm số cho lớp học và môn học này."
        )

    # Kiểm tra tồn tại của các tham chiếu cấu trúc
    cls = db.get(tables.Class, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail="Lớp học không tồn tại")
    subj = db.get(tables.Subject, subject_id)
    if not subj:
        raise HTTPException(status_code=404, detail="Môn học không tồn tại")
    sem = db.get(tables.Semester, semester_id)
    if not sem:
        raise HTTPException(status_code=404, detail="Học kỳ không tồn tại")

    # Lấy danh sách ID học sinh thuộc lớp để kiểm tra bảo mật chéo
    enroll_stmt = select(tables.Enrollment.student_id).where(
        tables.Enrollment.class_id == class_id, tables.Enrollment.is_active.is_(True)
    )
    enrolled_student_ids = set(db.execute(enroll_stmt).scalars().all())

    # Map ID học sinh sang thông tin để kiểm tra chéo
    students_stmt = select(tables.Student).where(
        tables.Student.id.in_(enrolled_student_ids), tables.Student.is_active.is_(True)
    )
    students_map = {s.id: s for s in db.execute(students_stmt).scalars().all()}

    # Tải workbook Excel dạng data_only=True để lấy giá trị số thay vì công thức Excel
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(
            status_code=400, detail="Đọc tệp tin thất bại. Vui lòng tải lên tệp định dạng Excel (.xlsx) hợp lệ."
        )

    # Validate dòng tiêu đề để chắc chắn file mẫu đúng định dạng
    header_row = [cell.value for cell in ws[1]]
    if len(header_row) < 3 or header_row[0] != "ID Học sinh (Ẩn)" or header_row[1] != "Mã học sinh":
        raise HTTPException(
            status_code=400,
            detail="Cấu trúc file Excel mẫu không khớp hoặc đã bị thay đổi cấu trúc cột. Vui lòng tải file mẫu chuẩn.",
        )

    errors = []
    preview_data = []
    valid_rows_count = 0

    # Phân tích từng hàng bắt từ hàng 2
    for r_idx in range(2, ws.max_row + 1):
        student_id_val = ws.cell(row=r_idx, column=1).value
        student_code_val = ws.cell(row=r_idx, column=2).value
        student_name_val = ws.cell(row=r_idx, column=3).value

        # Bỏ qua nếu là dòng trống hoàn toàn
        if not student_id_val and not student_code_val and not student_name_val:
            continue

        if not student_id_val:
            errors.append(
                {
                    "row": r_idx,
                    "student_code": str(student_code_val or ""),
                    "student_name": str(student_name_val or ""),
                    "column": "ID Học sinh",
                    "value_received": "",
                    "error_message": "Thiếu mã định danh học sinh kỹ thuật.",
                }
            )
            continue

        try:
            student_uuid = UUID(str(student_id_val).strip())
        except ValueError:
            errors.append(
                {
                    "row": r_idx,
                    "student_code": str(student_code_val or ""),
                    "student_name": str(student_name_val or ""),
                    "column": "ID Học sinh",
                    "value_received": str(student_id_val),
                    "error_message": "Mã định danh học sinh không đúng định dạng.",
                }
            )
            continue

        if student_uuid not in enrolled_student_ids:
            errors.append(
                {
                    "row": r_idx,
                    "student_code": str(student_code_val or ""),
                    "student_name": str(student_name_val or ""),
                    "column": "Học sinh",
                    "value_received": str(student_uuid),
                    "error_message": "Học sinh này không thuộc danh sách ghi danh của lớp học đang chọn.",
                }
            )
            continue

        student = students_map.get(student_uuid)
        student_name = student.full_name if student else str(student_name_val or "")
        student_code = student.student_code if student else str(student_code_val or "")

        row_scores = {}
        row_has_error = False

        # Đọc điểm số từ cột D (index 4) tương ứng cấu trúc SCORE_COLUMNS
        for c_offset, (cat, idx) in enumerate(SCORE_COLUMNS, 4):
            cell_val = ws.cell(row=r_idx, column=c_offset).value
            col_label = column_label(cat, idx)

            if cell_val is None or str(cell_val).strip() == "":
                row_scores[f"{cat.value}_{idx}"] = None
                continue

            try:
                # Xử lý nhập dấu phẩy của người Việt (Ví dụ: 8,5 -> 8.5)
                clean_val = str(cell_val).replace(",", ".").strip()
                val_float = float(clean_val)
            except ValueError:
                errors.append(
                    {
                        "row": r_idx,
                        "student_code": student_code,
                        "student_name": student_name,
                        "column": col_label,
                        "value_received": str(cell_val),
                        "error_message": "Điểm số nhập vào bắt buộc phải là số.",
                    }
                )
                row_has_error = True
                continue

            if val_float < 0 or val_float > 10:
                errors.append(
                    {
                        "row": r_idx,
                        "student_code": student_code,
                        "student_name": student_name,
                        "column": col_label,
                        "value_received": str(cell_val),
                        "error_message": "Điểm số không hợp lệ (Phải nằm trong thang điểm từ 0.0 đến 10.0).",
                    }
                )
                row_has_error = True
                continue

            # Làm tròn 2 số thập phân để khớp với DB Numeric(4, 2)
            row_scores[f"{cat.value}_{idx}"] = round(val_float, 2)

        if not row_has_error:
            valid_rows_count += 1
            preview_data.append(
                {
                    "student_id": str(student_uuid),
                    "student_code": student_code,
                    "student_name": student_name,
                    "scores": row_scores,
                }
            )

    return {
        "success": True,
        "total_rows": ws.max_row - 1,
        "valid_rows_count": valid_rows_count,
        "invalid_rows_count": len(errors),
        "errors": errors,
        "preview_data": preview_data,
    }


@router.post("/import/confirm")
def confirm_import(
    payload: ScoreImportConfirmRequest,
    db: Session = Depends(get_db),
    current_user: tables.User = Depends(get_current_user),
):
    """Xác nhận và ghi hàng loạt điểm số hợp lệ từ file Excel vào database."""
    # Kiểm tra quyền ghi điểm của user
    if not rbac.can_write_score(db, current_user, payload.subject_id, payload.class_id):
        raise HTTPException(
            status_code=403, detail="Bạn không có quyền nhập hoặc chỉnh sửa điểm số cho lớp học và môn học này."
        )

    # Phân nhóm: Bản ghi có điểm số -> Upsert, Bản ghi trống điểm -> Delete khỏi DB
    deletes = []
    updates = []

    for r in payload.records:
        if r.value is None:
            deletes.append(r)
        else:
            updates.append(r)

    try:
        # Thực hiện các thao tác ghi DB trong một Transaction duy nhất
        # 1. Xử lý các đầu điểm bị xóa (để trống) trong Excel
        if deletes:
            delete_conds = []
            for d in deletes:
                delete_conds.append(
                    and_(
                        tables.Score.student_id == d.student_id,
                        tables.Score.score_category == d.score_category,
                        tables.Score.column_index == d.column_index,
                    )
                )

            stmt = tables.Score.__table__.delete().where(
                and_(
                    tables.Score.subject_id == payload.subject_id,
                    tables.Score.semester_id == payload.semester_id,
                    or_(*delete_conds),
                )
            )
            db.execute(stmt)

        # 2. Xử lý Bulk Upsert dữ liệu điểm số mới/cập nhật
        if updates:
            insert_vals = [
                {
                    "student_id": u.student_id,
                    "subject_id": payload.subject_id,
                    "class_id": payload.class_id,
                    "semester_id": payload.semester_id,
                    "score_category": u.score_category,
                    "column_index": u.column_index,
                    "value": u.value,
                    "status": enums.ScoreStatus.DRAFT.value,
                    "entered_by": current_user.id,
                }
                for u in updates
            ]

            stmt = pg_insert(tables.Score).values(insert_vals)
            # Ràng buộc trùng khớp: student_id, subject_id, semester_id, category, index
            upsert_stmt = stmt.on_conflict_do_update(
                constraint="uq_score_unique",
                set_={
                    "value": stmt.excluded.value,
                    "status": stmt.excluded.status,
                    "entered_by": stmt.excluded.entered_by,
                    "updated_at": text("now()"),
                },
            )
            db.execute(upsert_stmt)

        db.commit()
        return {
            "success": True,
            "message": f"Ghi điểm thành công. Đã cập nhật {len(updates)} đầu điểm, xóa {len(deletes)} đầu điểm trống.",
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi hệ thống khi ghi cơ sở dữ liệu: {str(e)}")
